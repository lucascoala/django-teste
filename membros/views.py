from django.contrib.auth.decorators import permission_required
from django.contrib.auth.mixins import LoginRequiredMixin, PermissionRequiredMixin
from django.db.models import Count
from django.views.generic.edit import CreateView, UpdateView, DeleteView
from django.views.generic.list import ListView
from django.views.generic.detail import DetailView
from django.urls import reverse_lazy
from django.views.generic import TemplateView, View
from datetime import datetime
from django.template.loader import render_to_string
from django.http import HttpResponse
from weasyprint import HTML
import openpyxl
from openpyxl.styles import Font, Alignment
from braces.views import GroupRequiredMixin
from .models import Membros, Perfil
from django.shortcuts import redirect
from django.core.exceptions import PermissionDenied
from django.contrib import messages
from django.utils import timezone
from .forms import MembrosForm
from .utils import render_to_pdf


class MembrosCreate(GroupRequiredMixin, LoginRequiredMixin, CreateView):
    login_url = reverse_lazy('login')
    group_required = u"Administradores", u"Secretaria", u"Pastor"
    model = Membros
    form_class = MembrosForm
    template_name = 'membros/form.html'
    success_url = reverse_lazy('listar-membros')
    def dispatch(self, request, *args, **kwargs):
        if not request.user.is_authenticated:
            return redirect('login')
            
        if request.user.is_superuser:
            return super().dispatch(request, *args, **kwargs)
        
        # Verify module access
        if not hasattr(request.user, 'perfil'):
            perfil, created = Perfil.objects.get_or_create(usuario=request.user)
        
        if not request.user.perfil.acesso_membros:
            raise PermissionDenied("Você não tem permissão para acessar este módulo.")
            
        return super().dispatch(request, *args, **kwargs)

    def form_valid(self, form):
        # If creating a user with special role, verify permissions
        cargo = form.cleaned_data.get('cargo')
        special_roles = ['Pastor', 'Presbítero', 'Diácono', 'Secretaria', 'Tesoureiro']
        
        # Se não for superusuário e não estiver no grupo Administradores, verifica permissões
        if not self.request.user.is_superuser and not self.request.user.groups.filter(name='Administradores').exists():
            if cargo in special_roles:
                raise PermissionDenied("Você não tem permissão para criar membros com este cargo.")
            
        response = super().form_valid(form)
        messages.success(self.request, 'Membro cadastrado com sucesso!')
        return response

    def get_context_data(self, *args, **kwargs):
        context = super().get_context_data(*args, **kwargs)
        context['titulo'] = "Cadastro de membros"
        context['botao'] = "Cadastrar"
        return context

class MembrosUptade(GroupRequiredMixin, LoginRequiredMixin, UpdateView):
    login_url = reverse_lazy('login')
    group_required = u"Administradores", u"Secretaria", u"Pastor"
    model = Membros
    form_class = MembrosForm
    template_name = 'membros/form.html'
    success_url = reverse_lazy('listar-membros')

    def dispatch(self, request, *args, **kwargs):
        if not request.user.is_authenticated:
            return redirect('login')
            
        if request.user.is_superuser:
            return super().dispatch(request, *args, **kwargs)
        
        if not hasattr(request.user, 'perfil'):
            perfil, created = Perfil.objects.get_or_create(usuario=request.user)
        
        if not request.user.perfil.acesso_membros:
            raise PermissionDenied("Você não tem permissão para acessar este módulo.")
            
        return super().dispatch(request, *args, **kwargs)

    def get_context_data(self, *args, **kwargs):
        context = super().get_context_data(*args, **kwargs)
        context['titulo'] = "Editar membro"
        context['botao'] = "Salvar"
        
        return context
    
    def form_valid(self, form):
        response = super().form_valid(form)
        messages.success(self.request, 'Membro editado com sucesso!')
        return response


class MembrosDelete(GroupRequiredMixin, LoginRequiredMixin, DeleteView):
    login_url = reverse_lazy('login')
    group_required = u"Administradores", u"Secretaria"
    model = Membros
    template_name = 'membros/form-excluir.html'
    success_url = reverse_lazy('listar-membros')

class MembrosList(GroupRequiredMixin, LoginRequiredMixin, ListView):
    login_url = reverse_lazy('login')
    group_required = u"Administradores", u"Secretaria", u"Pastor", u"Presbítero", u"Diácono"
    model = Membros
    template_name = 'membros/listas/membros-list.html'
    paginate_by = 5
    def dispatch(self, request, *args, **kwargs):
        if not request.user.is_authenticated:
            return redirect('login')
            
        if request.user.is_superuser:
            return super().dispatch(request, *args, **kwargs)
            
        # Verify module access
        if not hasattr(request.user, 'perfil'):
            perfil, created = Perfil.objects.get_or_create(usuario=request.user)
        
        if not request.user.perfil.acesso_membros:
            raise PermissionDenied("Você não tem permissão para acessar este módulo.")
            
        return super().dispatch(request, *args, **kwargs)
    def get_queryset(self):
        queryset = super().get_queryset()
        
        # Filter based on user role
        user = self.request.user
        if not user.groups.filter(name__in=['Administradores', 'Pastor', 'Secretaria']).exists():
            # Regular members can only see active members
            queryset = queryset.filter(cargo='Membro Ativo')
            
        # Search filter
        txt_nome = self.request.GET.get('nome')
        if txt_nome:
            queryset = queryset.filter(nome__icontains=txt_nome)
            
        # Explicitly select only the fields used in the template to avoid querying unused fields
        queryset = queryset.only('nome', 'cargo', 'telefone', 'endereco', 'faixa_etaria', 'pk')
            
        return queryset.order_by('nome')
class MembrosView(GroupRequiredMixin, LoginRequiredMixin, DetailView):
    login_url = reverse_lazy('login')
    group_required = u"Administradores", u"Secretaria", u"Pastor"
    template_name = 'membros/visualizar.html'
    model = Membros
    context_object_name = 'membro'

class DashboardView(GroupRequiredMixin, LoginRequiredMixin, ListView):
    login_url = reverse_lazy('login')
    group_required = u"Administradores", u"Secretaria", u"Pastor", u"Presbítero", u"Diácono"
    model = Membros
    template_name = 'membros/dashboard.html'
    
    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)
        hoje = timezone.now()
        
        # Estatísticas gerais
        context['total_membros'] = Membros.objects.count()
        context['membros_ativos'] = Membros.objects.filter(cargo='Membro Ativo').count()
        context['total_lideres'] = Membros.objects.filter(
            cargo__in=['Pastor', 'Presbítero', 'Diácono']
        ).count()
        context['novos_membros'] = Membros.objects.filter(
            data_cadastro__month=hoje.month,
            data_cadastro__year=hoje.year
        ).count()
        
        # Aniversariantes do mês
        context['aniversariantes'] = Membros.objects.filter(
            data_nascimento__month=hoje.month
        ).order_by('data_nascimento__day')
        
        # Aniversários de casamento do mês
        context['aniversarios_casamento'] = Membros.objects.filter(
            data_casamento__month=hoje.month
        ).order_by('data_casamento__day')
        
        # Distribuição por faixa etária
        faixa_etaria = Membros.objects.values('faixa_etaria').annotate(
            total=Count('id')
        )
        context['faixa_etaria'] = {item['faixa_etaria']: item['total'] for item in faixa_etaria if item['faixa_etaria']}
        
        return context

@permission_required('membros.view_membro')
def exportar_membros_excel(request):
    # Criar um novo workbook e selecionar a planilha ativa
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Membros"

    # Definir cabeçalhos
    headers = [
        'Nome', 'Cargo', 'Estado Civil', 'Profissão', 
        'Escolaridade', 'Data Nascimento', 'Data Batismo',
        'Endereço', 'Número', 'Bairro', 'Cidade', 'Estado', 'CEP',
        'Telefone', 'Telefone Fixo', 'Email',
        'Data Casamento', 'Faixa Etária', 'Status'
    ]

    # Estilo para cabeçalhos
    header_font = Font(bold=True)
    header_alignment = Alignment(horizontal='center')

    # Adicionar cabeçalhos
    for col, header in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col)
        cell.value = header
        cell.font = header_font
        cell.alignment = header_alignment

    # Buscar todos os membros
    membros = Membros.objects.all().order_by('nome')

    # Adicionar dados dos membros
    for row, membro in enumerate(membros, 2):
        data = [
            membro.nome,
            membro.cargo,
            membro.estado_civil,
            membro.profissao,
            membro.escolaridade,
            membro.data_nascimento.strftime('%d/%m/%Y') if membro.data_nascimento else '',
            membro.data_batismo.strftime('%d/%m/%Y') if membro.data_batismo else '',
            membro.endereco,
            membro.numero,
            membro.bairro,
            membro.cidade,
            membro.estado,
            membro.cep,
            membro.telefone,
            membro.telefone_fixo,
            membro.email,
            membro.data_casamento.strftime('%d/%m/%Y') if membro.data_casamento else '',
            membro.faixa_etaria,
            'Ativo' if membro.is_active else 'Inativo'
        ]

        for col, value in enumerate(data, 1):
            cell = ws.cell(row=row, column=col)
            cell.value = value
            cell.alignment = Alignment(horizontal='left')

    # Ajustar largura das colunas
    for col in ws.columns:
        max_length = 0
        column = col[0].column_letter
        for cell in col:
            try:
                if len(str(cell.value)) > max_length:
                    max_length = len(str(cell.value))
            except:
                pass
        adjusted_width = (max_length + 2)
        ws.column_dimensions[column].width = adjusted_width

    # Criar a resposta HTTP com o arquivo Excel
    response = HttpResponse(content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
    response['Content-Disposition'] = f'attachment; filename=membros_{datetime.now().strftime("%d-%m-%Y")}.xlsx'

    # Salvar o arquivo
    wb.save(response)

    return response

class GerarPDFMembros(GroupRequiredMixin, LoginRequiredMixin, View):
    login_url = reverse_lazy('login')
    group_required = u"Administradores"

    def get(self, request, *args, **kwargs):
        # Dados para o PDF
        membros = Membros.objects.all()
        data = {
            'membros': membros,
            'data_atual': datetime.now().strftime("%d/%m/%Y"),
            'page_num': 1
        }
        
        # Gerar PDF
        pdf = render_to_pdf('membros/pdf_template.html', data)
        if pdf:
            response = HttpResponse(pdf, content_type='application/pdf')
            filename = "membros_%s.pdf" % datetime.now().strftime("%Y%m%d_%H%M%S")
            content = "inline; filename='%s'" % filename
            response['Content-Disposition'] = content
            return response
        return HttpResponse("Erro ao gerar PDF")
